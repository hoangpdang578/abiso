from openpyxl import load_workbook
from model.accounts import Accounts
from properties import USER_AVISO, ACCOUNTS_FILE_PATH


class Sheet:

    def __init__(self):
        self.accounts = []
        self.workbook = load_workbook(filename=ACCOUNTS_FILE_PATH)
        self.sheet = self.workbook.active

    def get_all_accounts(self):
        for row in self.sheet.iter_rows(min_row=2):
            if row[USER_AVISO].value is not None:
                account = Accounts(row)
                self.accounts.append(account)
        return self.accounts

    def update_account(self, account):
        self.sheet[account.get_id_no()] = account.get_row()
        self.workbook.save(filename=ACCOUNTS_FILE_PATH)
